"""schema-contract.md §5 — 세무사 전달용 XLSX. 청구 전체 반려(excluded=true)된
claim이 합계에서 빠지는지가 이 스위트의 핵심 계약이다 — 실제로 안 나간 돈이
합계에 남으면 세무사가 은행 내역과 대조할 때 어긋난다."""

from io import BytesIO

import pytest
from openpyxl import load_workbook

from src.settlements import export


def _claim(claim_id, **overrides):
    claim = {
        "claim_id": claim_id,
        "receipt_id": f"rct_{claim_id}",
        "recipient_id": "rcp_1",
        "currency": "KRW",
        "amount_minor": 10000,
        "account_category_code": "TRAVEL",
        "status": "SETTLED",
    }
    claim.update(overrides)
    return claim


@pytest.fixture(autouse=True)
def _wire(monkeypatch):
    monkeypatch.setattr(export, "get_settlement_run", lambda run_id: {"settlement_run_id": run_id})
    monkeypatch.setattr(export, "get_recipient", lambda rid: {"display_name": "박수현"})


def _rows(xlsx_bytes: bytes) -> list[tuple]:
    wb = load_workbook(BytesIO(xlsx_bytes))
    ws = wb.active
    return [tuple(row) for row in ws.iter_rows(min_row=2, values_only=True)]


def test_unknown_run_raises(monkeypatch):
    monkeypatch.setattr(export, "get_settlement_run", lambda run_id: None)
    with pytest.raises(export.RunNotFound):
        export.build_settlement_export("run_missing")


def test_excluded_claim_is_listed_but_not_counted_in_total(monkeypatch):
    monkeypatch.setattr(
        export,
        "get_claims_for_run",
        lambda run_id: [
            _claim("clm_1", amount_minor=10000),
            _claim("clm_2", amount_minor=9000, status="CONFIRMED", excluded=True),
        ],
    )
    monkeypatch.setattr(export, "_get_receipt", lambda receipt_id: None)

    rows = _rows(export.build_settlement_export("run_1"))

    clm_1_row = next(r for r in rows if r[0] == "clm_1")
    clm_2_row = next(r for r in rows if r[0] == "clm_2")
    assert clm_1_row[9] == "SETTLED"
    assert clm_2_row[9] == "CONFIRMED (Rejected)"

    total_row = next(r for r in rows if r[5] and r[5].startswith("Total"))
    assert total_row[8] == 10000  # clm_2(9000)는 합계에서 빠진다


def test_merchant_name_prefers_english_translation_and_category_is_english(monkeypatch):
    """해커톤 제출 언어 요건 — 엑셀 전체가 영어다. merchant_name_en이 있으면
    그걸 쓰고(가맹점명 한국어 원본 폴백은 번역 실패 시에만), 계정과목명도
    CATEGORY_DISPLAY_EN을 쓴다."""
    monkeypatch.setattr(
        export,
        "get_claims_for_run",
        lambda run_id: [_claim("clm_1", account_category_code="TRAVEL")],
    )
    monkeypatch.setattr(
        export,
        "_get_receipt",
        lambda receipt_id: {"merchant_name": "스타벅스", "merchant_name_en": "Starbucks"},
    )

    rows = _rows(export.build_settlement_export("run_1"))

    clm_1_row = next(r for r in rows if r[0] == "clm_1")
    assert clm_1_row[3] == "Starbucks"
    assert clm_1_row[5] == "Travel"


def test_merchant_name_falls_back_to_korean_when_not_translated(monkeypatch):
    monkeypatch.setattr(
        export, "get_claims_for_run", lambda run_id: [_claim("clm_1")]
    )
    monkeypatch.setattr(
        export, "_get_receipt", lambda receipt_id: {"merchant_name": "스타벅스"}
    )

    rows = _rows(export.build_settlement_export("run_1"))

    assert next(r for r in rows if r[0] == "clm_1")[3] == "스타벅스"


def test_all_claims_excluded_produces_no_total_row(monkeypatch):
    monkeypatch.setattr(
        export,
        "get_claims_for_run",
        lambda run_id: [_claim("clm_1", status="CONFIRMED", excluded=True)],
    )
    monkeypatch.setattr(export, "_get_receipt", lambda receipt_id: None)

    rows = _rows(export.build_settlement_export("run_1"))

    assert not any(r[5] and r[5].startswith("Total") for r in rows)
