"""schema-contract.md §5 — 정산 배치의 세무사 전달용 XLSX. 계정과목 컬럼 포함.

금액은 claims.amount_minor(정수)를 그대로 쓴다 — LLM이나 이 모듈이 새로 계산하지
않는다(money-safety.md 절대 규칙). 표시용 소수 문자열만 payouts/currency.py의
기존 변환 함수로 만든다.
"""

from io import BytesIO

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from ..payouts.currency import CURRENCY_EXPONENT, minor_to_paypal_value
from ..payouts.store import get_claims_for_run, get_client, get_recipient, get_settlement_run
from ..schemas.enums import CATEGORY_DISPLAY

_HEADERS = [
    "청구ID",
    "수취인",
    "거래일자",
    "가맹점명",
    "계정과목코드",
    "계정과목명",
    "통화",
    "금액",
    "금액(최소단위)",
    "상태",
]


class RunNotFound(Exception):
    pass


def _get_receipt(receipt_id: str) -> dict | None:
    doc = get_client().collection("receipts").document(receipt_id).get()
    return doc.to_dict() if doc.exists else None


def _amount_display(amount_minor: int, currency: str) -> str:
    if currency not in CURRENCY_EXPONENT:
        return str(amount_minor)
    return minor_to_paypal_value(amount_minor, currency)


def build_settlement_export(run_id: str) -> bytes:
    run = get_settlement_run(run_id)
    if run is None:
        raise RunNotFound(run_id)

    claims = get_claims_for_run(run_id)
    recipient_names: dict[str, str] = {}

    wb = Workbook()
    ws = wb.active
    ws.title = "정산내역"
    ws.append(_HEADERS)

    total_by_currency: dict[str, int] = {}
    for claim in sorted(claims, key=lambda c: c.get("claim_id", "")):
        recipient_id = claim["recipient_id"]
        if recipient_id not in recipient_names:
            recipient = get_recipient(recipient_id)
            recipient_names[recipient_id] = recipient["display_name"] if recipient else recipient_id

        receipt = _get_receipt(claim["receipt_id"])
        currency = claim["currency"]
        amount_minor = claim["amount_minor"]
        category_code = claim["account_category_code"]

        # excluded(청구 전체 반려, settlements/routes.py._apply_claim_exclusion)된
        # claim은 amount_minor를 그대로 갖고 있다(줄이지 않는다 — 되돌릴 때 복원할
        # 값이 있어야 하므로). 세무사용 합계에는 실제로 안 나간 돈이니 넣지 않는다 —
        # 행 자체는 감사 추적용으로 그대로 남기고 상태만 표시한다.
        status_display = f"{claim['status']} (반려됨)" if claim.get("excluded") else claim["status"]
        ws.append(
            [
                claim["claim_id"],
                recipient_names[recipient_id],
                receipt.get("transaction_date") if receipt else None,
                receipt.get("merchant_name") if receipt else None,
                category_code,
                CATEGORY_DISPLAY.get(category_code, category_code),
                currency,
                _amount_display(amount_minor, currency),
                amount_minor,
                status_display,
            ]
        )
        if not claim.get("excluded"):
            total_by_currency[currency] = total_by_currency.get(currency, 0) + amount_minor

    for currency, minor_sum in sorted(total_by_currency.items()):
        ws.append(
            [
                None,
                None,
                None,
                None,
                None,
                f"합계 ({currency})",
                currency,
                _amount_display(minor_sum, currency),
                minor_sum,
                None,
            ]
        )

    for i, header in enumerate(_HEADERS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = max(len(header) + 2, 12)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
